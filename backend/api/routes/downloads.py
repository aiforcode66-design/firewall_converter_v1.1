# -*- coding: utf-8 -*-
"""
Download API routes - for backward compatibility.
"""
import sys
import os

from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from fastapi.responses import StreamingResponse
from io import BytesIO

# Add parent directory to path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../../..')))

from backend.api.routes.conversions import router as conversions_router

router = APIRouter(prefix="/api", tags=["downloads"])


@router.post("/download")
async def download_conversion(
    source_vendor: str = Form(...),
    destination_vendor: str = Form(...),
    format: str = Form("txt"),
    config_file: UploadFile = File(None),
    checkpoint_objects: UploadFile = File(None),
    checkpoint_policy: UploadFile = File(None),
    checkpoint_nat: UploadFile = File(None),
    checkpoint_config: UploadFile = File(None),
    interface_mapping_data: str = Form("{}"),
    target_layout_data: str = Form("[]"),
    exclude_unused: str = Form("off"),
    **generator_options
):
    """
    Download converted configuration.
    This endpoint provides backward compatibility with the old Flask API.
    """
    # Import here to avoid circular dependency
    from backend.services.parser_service import ParserService
    from backend.services.conversion_service import ConversionService
    from backend.api.models.schemas import MappingData, GeneratorOptions
    import json
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    try:
        # Parse mapping data
        try:
            mapping_data_dict = eval(interface_mapping_data) if interface_mapping_data else {}
            target_layout_list = eval(target_layout_data) if target_layout_data else []
        except Exception:
            mapping_data_dict = {}
            target_layout_list = []

        # Build files dictionary
        files = {
            "config_file": config_file,
            "checkpoint_objects": checkpoint_objects,
            "checkpoint_policy": checkpoint_policy,
            "checkpoint_nat": checkpoint_nat,
            "checkpoint_config": checkpoint_config,
        }

        # Parse source configuration
        source_config = await ParserService.parse_config(source_vendor, files)

        # Perform conversion
        script, stats = ConversionService.convert_config(
            source_config=source_config,
            dest_vendor=destination_vendor,
            interface_mapping=mapping_data_dict.get("interface_mapping", {}),
            zone_mapping=mapping_data_dict.get("zone_mapping", {}),
            target_layout=target_layout_list,
            generator_options={k: v for k, v in generator_options.items() if v},
            exclude_unused=exclude_unused == "on",
        )

        if format == "txt":
            output = BytesIO(script.encode("utf-8"))
            filename = "conversion.txt"
            media_type = "text/plain"
        elif format == "xlsx":
            wb = Workbook()
            ws = wb.active
            ws.title = "Configuration"

            ws['A1'] = "Converted Configuration"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = script

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            filename = "conversion.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            raise HTTPException(status_code=400, detail="Invalid format")

        return StreamingResponse(
            output,
            media_type=media_type,
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Download failed: {str(e)}")
