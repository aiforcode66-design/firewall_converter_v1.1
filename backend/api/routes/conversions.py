# -*- coding: utf-8 -*-
"""
Conversion API routes.
"""
import sys
import os
import uuid
from datetime import datetime
from typing import Dict, Optional

from fastapi import APIRouter, UploadFile, File, Form, HTTPException, Depends
from fastapi.responses import StreamingResponse
from fastapi import status as http_status
from io import BytesIO
from sqlalchemy.ext.asyncio import AsyncSession

# Add parent directory to path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../../..')))

from backend.api.models.schemas import (
    Vendor,
    ConversionResults,
    ConversionStatus,
    ConfigStats,
    ComparisonItem,
    MappingData,
    GeneratorOptions,
)
from backend.api.models.database import get_db, ConversionRepository
from backend.services.parser_service import ParserService
from backend.services.conversion_service import ConversionService

router = APIRouter(prefix="/api/v1/conversions", tags=["conversions"])

# In-memory cache for conversion scripts (scripts stored in memory, metadata in DB)
conversion_cache: Dict[str, dict] = {}


@router.post("", response_model=ConversionResults)
async def create_conversion(
    source_vendor: str = Form(...),
    destination_vendor: str = Form(...),
    config_file: Optional[UploadFile] = File(None),
    checkpoint_objects: Optional[UploadFile] = File(None),
    checkpoint_policy: Optional[UploadFile] = File(None),
    checkpoint_nat: Optional[UploadFile] = File(None),
    checkpoint_config: Optional[UploadFile] = File(None),
    checkpoint_csv_zip: Optional[UploadFile] = File(None),
    interface_mapping_data: str = Form("{}"),
    target_layout_data: str = Form("[]"),
    target_ip_config: str = Form("{}"),
    exclude_unused: str = Form("off"),
    generator_options: str = Form("{}"),
    db: AsyncSession = Depends(get_db),
):
    """
    Create a new conversion job and save to database history.
    """
    try:
        # Validate vendors
        try:
            src_vendor = Vendor(source_vendor)
            dst_vendor = Vendor(destination_vendor)
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid vendor. Supported: {[v.value for v in Vendor]}"
            )

        import json

        # Parse mapping data
        try:
            mapping_data_dict = json.loads(interface_mapping_data) if interface_mapping_data else {}
            target_layout_list = json.loads(target_layout_data) if target_layout_data else []
            ip_config = json.loads(target_ip_config) if target_ip_config else {}
        except Exception:
            mapping_data_dict = {}
            target_layout_list = []
            ip_config = {}

        mapping = MappingData(
            interface_mapping=mapping_data_dict.get("interface_mapping", {}),
            zone_mapping=mapping_data_dict.get("zone_mapping", {}),
        )

        # Build generator options
        try:
            gen_options_dict = json.loads(generator_options) if generator_options else {}
        except:
            gen_options_dict = {}
        gen_options = GeneratorOptions(**{k: v for k, v in gen_options_dict.items() if v})

        # Build files dictionary
        files: Dict[str, Optional[UploadFile]] = {
            "config_file": config_file,
            "checkpoint_objects": checkpoint_objects,
            "checkpoint_policy": checkpoint_policy,
            "checkpoint_nat": checkpoint_nat,
            "checkpoint_config": checkpoint_config,
            "checkpoint_csv_zip": checkpoint_csv_zip,
        }

        # Parse source configuration
        source_config = await ParserService.parse_config(source_vendor, files)

        # Perform conversion
        script, stats = ConversionService.convert_config(
            source_config=source_config,
            dest_vendor=destination_vendor,
            interface_mapping=mapping.interface_mapping,
            zone_mapping=mapping.zone_mapping,
            target_layout=target_layout_list,
            generator_options=gen_options.dict(),
            exclude_unused=exclude_unused == "on",
            source_vendor=source_vendor,
        )

        # Build tables
        comparison_table = ConversionService.build_comparison_table(source_config, destination_vendor)
        objects_table = ConversionService.build_objects_table(source_config)
        rules_table = ConversionService.build_rules_table(source_config)
        nat_rules_table = ConversionService.build_nat_rules_table(source_config)
        routes_table = ConversionService.build_routes_table(source_config)
        time_ranges_table = ConversionService.build_time_ranges_table(source_config)
        conversion_warnings = ConversionService.build_conversion_warnings(source_config)

        # Create conversion result
        conversion_id = str(uuid.uuid4())
        file_name = config_file.filename if config_file else f"{source_vendor}_config"

        result = ConversionResults(
            conversion_id=conversion_id,
            config_id=str(uuid.uuid4()),
            source_vendor=src_vendor,
            dest_vendor=dst_vendor,
            status=ConversionStatus.COMPLETED,
            script=script,
            stats=ConfigStats(**stats),
            comparison_table=[ComparisonItem(**c) for c in comparison_table],
            objects_table=objects_table,
            rules_table=rules_table,
            nat_rules_table=nat_rules_table,
            routes_table=routes_table,
            time_ranges_table=time_ranges_table,
            conversion_warnings=conversion_warnings,
            created_at=datetime.utcnow().isoformat(),
        )

        # Store in cache and database
        conversion_cache[conversion_id] = {
            "result": result,
            "source_vendor": source_vendor,
            "dest_vendor": destination_vendor,
            "mapping": mapping.dict(),
            "target_layout": target_layout_list,
            "generator_options": gen_options.dict(),
            "exclude_unused": exclude_unused == "on",
        }

        # Save to database
        try:
            await ConversionRepository.create(
                db=db,
                id=conversion_id,
                source_vendor=source_vendor,
                dest_vendor=destination_vendor,
                file_name=file_name,
            )
            await ConversionRepository.update_status(
                db=db,
                id=conversion_id,
                status="completed",
                stats=stats,
                warnings=conversion_warnings,
            )
        except Exception as db_error:
            # Log but don't fail the conversion if DB save fails
            import traceback
            traceback.print_exc()
            print(f"Warning: Failed to save to database: {db_error}")

        return result

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Conversion failed: {str(e)}")


@router.get("/{conversion_id}", response_model=ConversionResults)
async def get_conversion(conversion_id: str):
    """Get conversion results by ID."""
    if conversion_id not in conversion_cache:
        raise HTTPException(status_code=http_status.HTTP_404_NOT_FOUND, detail="Conversion not found")
    return conversion_cache[conversion_id]["result"]


@router.post("/{conversion_id}/export/{format}")
async def export_conversion(conversion_id: str, format: str):
    """Export conversion result in specified format."""
    if conversion_id not in conversion_cache:
        raise HTTPException(status_code=http_status.HTTP_404_NOT_FOUND, detail="Conversion not found")

    session = conversion_cache[conversion_id]
    script = session["result"].script

    if format == "txt":
        # Export as text file
        output = BytesIO(script.encode("utf-8"))
        filename = f"conversion_{conversion_id[:8]}.txt"
        media_type = "text/plain"
    elif format == "xlsx":
        # Export as Excel (using openpyxl)
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Configuration"

            # Add script
            ws['A1'] = "Converted Configuration"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = script

            # Add rules table
            if session["result"].rules_table:
                ws_rules = wb.create_sheet(title="Rules")
                headers = list(session["result"].rules_table[0].keys())
                for col, header in enumerate(headers, 1):
                    cell = ws_rules.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")

                for row_idx, row in enumerate(session["result"].rules_table, 2):
                    for col_idx, (key, value) in enumerate(row.items(), 1):
                        ws_rules.cell(row=row_idx, column=col_idx, value=str(value))

            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            filename = f"conversion_{conversion_id[:8]}.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Excel export failed: {str(e)}")
    else:
        raise HTTPException(status_code=400, detail="Invalid format. Use 'txt' or 'xlsx'")

    return StreamingResponse(
        output,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
